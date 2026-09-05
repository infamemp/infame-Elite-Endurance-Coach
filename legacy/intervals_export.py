"""
intervals.icu Coach Athlete Exporter | PRODUCTION VERSION 1.6 (ULTIMATE)
Extracts data from all athletes and generates a multi-sheet Excel report.
"""

import os
import sys
import re
import base64
from datetime import datetime, timedelta, date

try:
    import requests
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    print("🚀 Loading PRODUCTION VERSION 1.6 (ULTIMATE)...")
except ImportError as e:
    print(f"❌ Critical error: Missing dependencies. {e}")
    sys.exit(1)

# ─────────────────────────────────────────────
API_KEY = os.getenv("ICU_API_KEY")
if not API_KEY:
    sys.exit("❌ Falta la variable de entorno ICU_API_KEY (ejecuta: setx ICU_API_KEY \"tu_key\")")
OUTPUT_FILE = "athletes_report_production.xlsx"
# ─────────────────────────────────────────────

BASE_URL = "https://intervals.icu/api/v1"

# ── Excel Styles ──────────────────────────────
H1_FILL   = PatternFill("solid", start_color="1F4E79")
H2_FILL   = PatternFill("solid", start_color="2E75B6")
H3_FILL   = PatternFill("solid", start_color="4472C4")
ALT_FILL  = PatternFill("solid", start_color="D6E4F0")
WHT_FILL  = PatternFill("solid", start_color="FFFFFF")
H1_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
H2_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
H3_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=9)
BODY      = Font(name="Arial", size=9)
BOLD      = Font(name="Arial", bold=True, size=9)
CTR       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT       = Alignment(horizontal="left",   vertical="center")
THIN      = Side(style="thin", color="AAAAAA")
BRD       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SESSION = None

def make_session():
    global SESSION
    s = requests.Session()
    token = base64.b64encode(f"API_KEY:{API_KEY}".encode()).decode()
    s.headers.update({"Authorization": f"Basic {token}", "Accept": "application/json"})
    return s

def get(endpoint, params=None):
    r = SESSION.get(f"{BASE_URL}{endpoint}", params=params, timeout=30)
    r.raise_for_status()
    return r.json()

def hrow(ws, row, ncols, fill=H3_FILL, font=H3_FONT):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = CTR
        cell.border = BRD

def drow(ws, row, ncols, alt=False):
    fill = ALT_FILL if alt else WHT_FILL
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = BODY
        cell.alignment = LFT
        cell.border = BRD

def autofit(ws, mn=8, mx=40):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=mn)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, mn), mx)

def v(val, default="—"):
    return val if val is not None and val != "" else default

def clean_n(val):
    try: 
        return float(val)
    except Exception: 
        return 0.0

def fmt_sec(secs):
    if not secs: return "—"
    h, r = divmod(int(secs), 3600)
    m, s = divmod(r, 60)
    return f"{h}h {m:02d}m" if h else f"{m}m {s:02d}s"

def calc_age(dob_str):
    if not dob_str: return "—"
    try:
        dob = datetime.strptime(dob_str, "%Y-%m-%d").date()
        today = date.today()
        return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))
    except Exception: 
        return "—"

def eftp_by_cat(summary_row):
    result = {}
    for bc in (summary_row.get("byCategory") or []):
        cat = bc.get("category","")
        result[cat] = {
            "eftp": bc.get("eftp"),
            "eftpPerKg": bc.get("eftpPerKg"),
        }
    return result

def sport_label(cat):
    return {"Ride": "Ride", "Run": "Run", "Swim": "Swim"}.get(cat, cat)

def fmt_pace(speed_ms, units="mins_km"):
    if not speed_ms or speed_ms == "—": return "—"
    try:
        s_ms = float(speed_ms)
        if s_ms <= 0: return "—"
        
        is_100m = "100m" in str(units).lower()
        secs = 100 / s_ms if is_100m else 1000 / s_ms
        m, s = divmod(int(round(secs)), 60)
        
        suffix = '/ 100m' if is_100m else '/ km'
        return f"{m}:{s:02d} {suffix}"
    except Exception: 
        return str(speed_ms)

# ─── GLOBAL SUMMARY SHEET ───────────────────────────────────────────────────

def build_summary(wb, summary_list, profiles_map, last_activities):
    ws = wb.create_sheet("Athletes Summary", 0)
    ws.freeze_panes = "A3"

    cols = [
        "Name", "ID", "Age", "Gender", "Country", "Weight (kg)",
        "CTL Fitness", "ATL Fatigue", "TSB Form", "Ramp Rate",
        "Ride eFTP (W)", "Ride W/kg",
        "Run eFTP (W)", "Run W/kg",
        "Swim eFTP (W)", "Resting HR", "Last Activity"
    ]

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    tc = ws["A1"]
    tc.value = f"Athletes Summary — {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    tc.fill = H1_FILL
    tc.font = H1_FONT
    tc.alignment = CTR

    for c, h in enumerate(cols, 1):
        ws.cell(row=2, column=c, value=h)
    hrow(ws, 2, len(cols), H2_FILL, H2_FONT)

    for i, s in enumerate(summary_list):
        aid   = s.get("athlete_id")
        name  = s.get("athlete_name", "—")
        prof  = profiles_map.get(aid, {})
        cats  = eftp_by_cat(s)

        f  = clean_n(s.get('fitness'))
        a  = clean_n(s.get('fatigue'))
        fm = clean_n(s.get('form'))
        rr = clean_n(s.get('rampRate'))

        r_eftp = round(cats.get("Ride", {}).get("eftp", 0) or 0, 1) or "—"
        r_wkg  = round(cats.get("Ride", {}).get("eftpPerKg", 0) or 0, 3) or "—"
        rn_eftp= round(cats.get("Run", {}).get("eftp", 0) or 0, 1) or "—"
        rn_wkg = round(cats.get("Run", {}).get("eftpPerKg", 0) or 0, 3) or "—"
        sw_eftp= round(cats.get("Swim", {}).get("eftp", 0) or 0, 1) or "—"

        row = [
            name, aid, calc_age(prof.get("icu_date_of_birth")), 
            v(prof.get("sex")), v(prof.get("country")), 
            v(prof.get("icu_weight") or prof.get("weight")),
            f"{f:.1f}", f"{a:.1f}", f"{fm:.1f}", f"{rr:.2f}",
            r_eftp, r_wkg, rn_eftp, rn_wkg, sw_eftp,
            v(prof.get("icu_resting_hr")), 
            v(last_activities.get(aid, "—"))
        ]

        r = i + 3
        for c, val in enumerate(row, 1): 
            ws.cell(row=r, column=c, value=val)
        drow(ws, r, len(cols), alt=(i % 2 == 1))

    autofit(ws)
    ws.column_dimensions["A"].width = 24
    print(f"  ✓ Summary Table generated.")


# ─── INDIVIDUAL ATHLETE SHEET ────────────────────────────────────────────────

def build_athlete_sheet(wb, aid, name, summary_row, master_profile):
    safe_name = re.sub(r'[\\/*?:\[\]]', '-', name or f"Athlete {aid}")
    safe = safe_name[:31].strip() or f"A-{aid}"
    ws = wb.create_sheet(safe)

    # ── Fetch Live Profile ──
    try:
        athlete_full = get(f"/athlete/{aid}")
        sport_settings_list = athlete_full.get("sportSettings", [])
        if isinstance(sport_settings_list, dict):
            sport_settings_list = [sport_settings_list]
    except Exception:
        athlete_full = master_profile
        sport_settings_list = []

    s_map = {}
    for sp in sport_settings_list:
        for t in sp.get("types", []):
            s_map[t] = sp

    row = 1

    def section_title(title, ncols=8, fill=H2_FILL, font=H2_FONT):
        nonlocal row
        ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
        c = ws.cell(row=row, column=1, value=title)
        c.fill = fill
        c.font = font
        c.alignment = CTR
        ws.row_dimensions[row].height = 18
        row += 1

    def kv(label, val):
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = BOLD
        ws.cell(row=row, column=2, value=v(val)).font = BODY
        ws.cell(row=row, column=1).border = BRD
        ws.cell(row=row, column=2).border = BRD
        ws.row_dimensions[row].height = 15
        row += 1

    # ── Personal Info ──
    section_title("PERSONAL INFORMATION", fill=H1_FILL, font=H1_FONT)
    
    nm = athlete_full.get("name") or master_profile.get("name") or name
    dob = athlete_full.get("icu_date_of_birth") or master_profile.get("icu_date_of_birth")
    sx = athlete_full.get("sex") or master_profile.get("sex")
    wt = athlete_full.get("icu_weight") or athlete_full.get("weight") or master_profile.get("icu_weight") or master_profile.get("weight")
    ht = athlete_full.get("height") or master_profile.get("height")
    ct = athlete_full.get("city") or master_profile.get("city")
    cy = athlete_full.get("country") or master_profile.get("country")
    
    kv("Name", nm)
    kv("Age", calc_age(dob))
    kv("Gender", sx)
    kv("Weight (kg)", wt)
    kv("Height (m)", ht)
    kv("City", ct)
    kv("Country", cy)
    
    r_hr_val = athlete_full.get("resting_hr") or athlete_full.get("icu_resting_hr") or master_profile.get("icu_resting_hr") or "—"
    kv("Resting HR", f"{r_hr_val} bpm" if r_hr_val != "—" else "—")
    row += 1

    # ── PMC — fuente: wellness endpoint (misma fuente que build_state.py) ──
    # No usar summary_row para PMC: el endpoint athlete-summary puede devolver
    # filas duplicadas con valores distintos para el mismo día, y la deduplicación
    # solo garantiza unicidad de athlete_id, no el valor más reciente.
    section_title("FITNESS · FATIGUE · FORM (PMC)", fill=H1_FILL, font=H1_FONT)

    today_str = date.today().isoformat()
    w_rec = None
    try:
        # Intentar el día de hoy; caer a ayer si el registro aún no existe
        for oldest in [today_str, (date.today() - timedelta(days=1)).isoformat()]:
            recs = get(f"/athlete/{aid}/wellness",
                       params={"oldest": oldest, "newest": today_str})
            if recs:
                w_rec = recs[-1]  # el más reciente dentro del rango
                break
    except Exception:
        pass  # si el endpoint falla, caemos al fallback

    if w_rec:
        f  = clean_n(w_rec.get("ctl"))
        a  = clean_n(w_rec.get("atl"))
        tsb_raw = w_rec.get("tsb")
        fm = (f - a) if tsb_raw is None else clean_n(tsb_raw)
        rr = clean_n(w_rec.get("rampRate"))
    else:
        # Fallback: valores del athlete-summary (pueden diferir levemente)
        f  = clean_n(summary_row.get("fitness"))
        a  = clean_n(summary_row.get("fatigue"))
        fm = clean_n(summary_row.get("form"))
        rr = clean_n(summary_row.get("rampRate"))
    
    kv("CTL – Fitness", f"{f:.1f}")
    kv("ATL – Fatigue", f"{a:.1f}")
    kv("TSB – Form", f"{fm:.1f}")
    kv("Ramp Rate", f"{rr:.2f}")
    kv("Weekly Load", v(summary_row.get("training_load")))
    kv("PMC Date", date.today().isoformat())
    row += 1

    # ── Sports Metrics (RESTORED TO CORRECT LOGIC USING s_map) ──
    section_title("SPORT CONFIGURATION & SETTINGS", fill=H1_FILL, font=H1_FONT)
    sport_hdrs = ["Sport", "Metric / Parameter", "Configured Value"]
    for c, h in enumerate(sport_hdrs, 1): 
        ws.cell(row=row, column=c, value=h)
    hrow(ws, row, len(sport_hdrs))
    row += 1

    summary_cats = eftp_by_cat(summary_row)
    target_sports = ["Ride", "Run", "Swim"]
    alt_color = False

    for sport in target_sports:
        sp = s_map.get(sport, {})
        sc = summary_cats.get(sport, {})

        if not sp and not sc.get("eftp"): 
            continue

        metrics_block = []

        if sport in ["Ride", "Run"]:
            mmp = sp.get("mmp_model") or sp.get("mmpModel") or {}
            
            ftp_val    = sp.get('ftp') or sp.get('ftp_value')
            indoor_ftp = sp.get('indoor_ftp') or sp.get('indoorFtp')
            w_prime    = sp.get('w_prime') or sp.get('wPrime')
            p_max      = sp.get('p_max') or sp.get('pMax')
            
            ew_prime   = sp.get('eftp_w_prime') or sp.get('eftpWPrime') or mmp.get('w_prime') or mmp.get('wPrime') or sc.get('eftp_w_prime') or sc.get('eftpWPrime') or sc.get('wPrime')
            ep_max     = sp.get('eftp_p_max') or sp.get('eftpPMax') or mmp.get('p_max') or mmp.get('pMax') or sc.get('eftp_p_max') or sc.get('eftpPMax') or sc.get('pMax')
            eftp_v     = sc.get('eftp') or mmp.get('ftp') or ftp_val

            metrics_block.extend([
                ("FTP", f"{ftp_val} W" if ftp_val else "—"),
                ("Indoor FTP", f"{indoor_ftp} W" if indoor_ftp else "—"),
                ("W' J", f"{w_prime} J" if w_prime else "—"),
                ("Pmax", f"{p_max} W" if p_max else "—"),
                ("eFTP", f"{round(eftp_v, 1)} W" if eftp_v else "—"),
                ("eW'", f"{ew_prime} J" if ew_prime else "—"),
                ("ePmax", f"{ep_max} W" if ep_max else "—"),
            ])
            
        elif sport == "Swim" and (sc.get("eftp") or sp.get("ftp")):
            swim_ftp = sc.get("eftp") or sp.get("ftp")
            metrics_block.append(("eFTP", f"{round(swim_ftp, 1)} W"))

        lt_hr = sp.get('lthr') or sp.get('ltHr') or sp.get('threshold_hr') or sp.get('thresholdHr')
        max_hr = sp.get('max_hr') or sp.get('maxHr')
        
        hrrc_val = sp.get('hrrc_min_hr') or sp.get('hrrcMinHr') or sp.get('hrrc_threshold') or sp.get('hrrc_min') or sp.get('hrrc')
        if not hrrc_val and lt_hr: 
            hrrc_val = lt_hr

        metrics_block.extend([
            ("Threshold HR", f"{lt_hr} bpm" if lt_hr else "—"),
            ("Max HR", f"{max_hr} bpm" if max_hr else "—"),
            ("HRRc", f"{hrrc_val} bpm" if hrrc_val else "—"),
            ("Min HR", f"{r_hr_val} bpm" if r_hr_val and r_hr_val != "—" else "—")
        ])

        if sport in ["Run", "Swim"]:
            lt_pace = sp.get("threshold_pace") or sp.get("lt_pace") or sp.get("ltPace") or sp.get("thresholdPace")
            pace_units = sp.get("pace_units") or sp.get("paceUnits") or "mins_km"
            metrics_block.extend([
                ("Threshold Pace", fmt_pace(lt_pace, pace_units)),
                ("Units", str(pace_units).upper())
            ])

        if not any(val != "—" for l, val in metrics_block if l not in ["Units", "Min HR"]): 
            continue

        first_row_sport = True
        for label, val in metrics_block:
            ws.cell(row=row, column=1, value=sport_label(sport) if first_row_sport else "")
            ws.cell(row=row, column=2, value=label)
            ws.cell(row=row, column=3, value=val)
            drow(ws, row, len(sport_hdrs), alt=alt_color)
            first_row_sport = False
            row += 1
            
        alt_color = not alt_color
        row += 1

    # ── Future Events & Races ──
    try:
        today  = date.today().isoformat()
        future = (date.today() + timedelta(days=365)).isoformat()
        ev_params = {"oldest": today, "newest": future}
        events = get(f"/athlete/{aid}/events", params=ev_params) or []
    except Exception:
        events = []

    races_list = []
    workouts_list = []

    for ev in events:
        category = str(ev.get("category") or "").upper()
        ev_type  = str(ev.get("type") or "").upper()
        priority = str(ev.get("priority") or "").upper()
        
        if "RACE" in category or "RACE" in ev_type or priority in ["A", "B", "C"]:
            races_list.append(ev)
        else:
            workouts_list.append(ev)

    # TABLE A: Races
    section_title("SCHEDULED RACES & COMPETITIONS (Next 365 days)", fill=H1_FILL, font=H1_FONT)
    evt_hdrs = ["Date", "Event Name", "Type", "Category", "Distance", "Priority"]
    for c, h in enumerate(evt_hdrs, 1): 
        ws.cell(row=row, column=c, value=h)
    hrow(ws, row, len(evt_hdrs))
    row += 1

    for i, ev in enumerate(races_list[:40]):
        cat_raw = str(ev.get("category") or "")
        priority = str(ev.get("priority") or "")
        
        if not priority and cat_raw.upper().startswith("RACE_"):
            priority = cat_raw.split("_")[-1]
            
        is_race = cat_raw.upper().startswith("RACE")
        display_cat = "Race" if is_race else cat_raw.capitalize()
        dist = ev.get("distance")
        
        ev_date = (ev.get("start_date_local") or "")[:10]
        
        ws.cell(row=row, column=1, value=ev_date)
        ws.cell(row=row, column=2, value=v(ev.get("name")))
        ws.cell(row=row, column=3, value=v(ev.get("type")))
        ws.cell(row=row, column=4, value=v(display_cat))
        ws.cell(row=row, column=5, value=f"{dist/1000:.1f} km" if dist else "—")
        ws.cell(row=row, column=6, value=v(priority))
        drow(ws, row, len(evt_hdrs), alt=(i % 2 == 1))
        row += 1
        
    if not races_list:
        ws.cell(row=row, column=1, value="No races or competitions scheduled").font = BODY
        row += 1
    row += 1

    # TABLE B: Workouts
    section_title("WORKOUTS, PLANS & REST DAYS SCHEDULE", fill=H3_FILL, font=H3_FONT)
    work_hdrs = ["Date", "Description / Session", "Type", "Est. Load", "Est. Time"]
    for c, h in enumerate(work_hdrs, 1): 
        ws.cell(row=row, column=c, value=h)
    hrow(ws, row, len(work_hdrs), fill=H3_FILL, font=H3_FONT)
    row += 1

    for i, ev in enumerate(workouts_list[:40]):
        ev_date = (ev.get("start_date_local") or "")[:10]
        ev_name = ev.get("name") or ev.get("description")
        ev_type = ev.get("type") or "Workout"
        
        plan_load = ev.get("icu_training_load") or ev.get("training_load") or ev.get("plan_load")
        d1 = ev.get("moving_time")
        d2 = ev.get("elapsed_time")
        d3 = ev.get("target_duration")
        d4 = ev.get("plan_duration")
        d5 = ev.get("duration")
        plan_time = d1 or d2 or d3 or d4 or d5

        ws.cell(row=row, column=1, value=ev_date)
        ws.cell(row=row, column=2, value=v(ev_name))
        ws.cell(row=row, column=3, value=v(ev_type))
        ws.cell(row=row, column=4, value=round(plan_load) if plan_load else "—")
        ws.cell(row=row, column=5, value=fmt_sec(plan_time))
        
        drow(ws, row, len(work_hdrs), alt=(i % 2 == 1))
        row += 1
        
    if not workouts_list:
        ws.cell(row=row, column=1, value="No planned workouts in schedule").font = BODY
        row += 1
    row += 1

    # ── History ──
    section_title("HISTORY: LAST 4 WEEKS ACTIVITIES", fill=H1_FILL, font=H1_FONT)
    act_hdrs = [
        "Date", "Name", "Sport", "Duration", "Dist (km)",
        "TSS/Load", "IF", "Avg Power (W)", "Avg HR",
        "Elevation (m)", "Avg Speed"
    ]
    for c, h in enumerate(act_hdrs, 1): 
        ws.cell(row=row, column=c, value=h)
    hrow(ws, row, len(act_hdrs))
    row += 1

    last_act_date = "—"
    try:
        oldest = (date.today() - timedelta(days=28)).isoformat()
        newest = date.today().isoformat()
        act_params = {"oldest": oldest, "newest": newest}
        acts = get(f"/athlete/{aid}/activities", params=act_params)
        acts = sorted(acts or [], key=lambda x: x.get("start_date_local",""), reverse=True)

        if acts:
            last_act_date = (acts[0].get("start_date_local") or "")[:10]

        for i, act in enumerate(acts):
            dist = act.get("distance")
            spd  = act.get("average_speed")
            intf = act.get("icu_intensity")
            act_date = (act.get("start_date_local") or "")[:10]
            mov_time = act.get("moving_time") or act.get("elapsed_time")
            t_load = act.get("icu_training_load")
            if t_load is None:
                t_load = act.get("training_load")
            # Avg Power: icu_weighted_avg_watts (NP calculado por ICU),
            # luego average_watts como fallback para medidores externos
            avg_w = (act.get("icu_weighted_avg_watts") or
                     act.get("average_watts"))
            
            ws.cell(row=row, column=1,  value=act_date)
            ws.cell(row=row, column=2,  value=v(act.get("name")))
            ws.cell(row=row, column=3,  value=v(act.get("type")))
            ws.cell(row=row, column=4,  value=fmt_sec(mov_time))
            ws.cell(row=row, column=5,  value=round(dist/1000, 2) if dist else "—")
            ws.cell(row=row, column=6,  value=f"{float(t_load):.2f}" if t_load is not None else "—")
            ws.cell(row=row, column=7,  value=f"{intf/100:.2f}" if intf is not None else "—")
            ws.cell(row=row, column=8,  value=v(avg_w))
            ws.cell(row=row, column=9,  value=v(act.get("average_heartrate")))
            elev = act.get("total_elevation_gain")
            ws.cell(row=row, column=10, value=round(elev) if elev else "—")
            ws.cell(row=row, column=11, value=f"{spd*3.6:.1f} km/h" if spd else "—")
            
            drow(ws, row, len(act_hdrs), alt=(i % 2 == 1))
            row += 1

        if not acts: 
            ws.cell(row=row, column=1, value="No activities in the last 4 weeks").font = BODY
            
    except Exception as e:
        ws.cell(row=row, column=1, value=f"Error loading activities: {e}").font = BODY

    autofit(ws)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.freeze_panes = "A2"
    print(f"  ✓ Processed: {safe}")
    
    return last_act_date

# ─── MAIN ────────────────────────────────────────────────────────────────────

def main():
    global SESSION
    print("🔐 Connecting to Intervals.icu API...")
    if API_KEY in ("API_Key_HERE", "TU_API_KEY_AQUI"):
        print("❌ Error: Set API_KEY.")
        sys.exit(1)

    SESSION = make_session()

    try:
        summary_list = get("/athlete/0/athlete-summary.json")
    except Exception as e:
        print(f"❌ Handshake error: {e}")
        sys.exit(1)

    if not summary_list:
        print("❌ No athletes found.")
        sys.exit(1)

    # Fusionar duplicados por athlete_id — el API devuelve hasta 2 filas por atleta.
    # Se toma como primaria la fila con mayor fitness (más reciente); la secundaria
    # rellena solo los campos que el primario tenga en None. Sin warnings: es
    # comportamiento esperado del endpoint.
    seen: dict = {}
    for s in summary_list:
        aid = s.get("athlete_id")
        if not aid:
            continue
        if aid not in seen:
            seen[aid] = dict(s)
        else:
            existing = seen[aid]
            f_new      = s.get("fitness") or 0
            f_existing = existing.get("fitness") or 0
            primary, secondary = (s, existing) if f_new > f_existing else (existing, s)
            merged = dict(secondary)
            merged.update({k: v for k, v in primary.items() if v is not None})
            seen[aid] = merged
    summary_list = list(seen.values())
    print(f"   ✓ {len(summary_list)} unique athletes")

    print("📥 Indexing profiles...")
    try: 
        profiles_map = {p["id"]: p for p in get("/athletes")}
    except Exception: 
        profiles_map = {}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("\n📋 Generating individual sheets...")
    last_activities = {}
    for s in summary_list:
        aid = s.get("athlete_id")
        try:
            last_activities[aid] = build_athlete_sheet(wb, aid, s.get("athlete_name"), s, profiles_map.get(aid, {}))
        except Exception as e:
            print(f"⚠️  Skipping {aid}: {e}")
            last_activities[aid] = "—"

    print("\n📊 Generating Global Summary...")
    build_summary(wb, summary_list, profiles_map, last_activities)

    out = os.path.abspath(OUTPUT_FILE)
    try:
        wb.save(out)
        print(f"\n✅ Done! Report saved to:\n   {out}")
    except IOError:
        print(f"\n❌ Error: Close '{OUTPUT_FILE}' before running.")

print("▶️ Code loaded. Starting execution...")
if __name__ == "__main__":
    main()